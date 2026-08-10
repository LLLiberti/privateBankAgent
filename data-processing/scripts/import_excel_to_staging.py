"""Export four-dimensional entrepreneur Excel workbooks to MySQL staging artifacts.

The script is deliberately database-free: it never opens a MySQL connection.
It reads the source workbooks and produces CSV/JSON artifacts plus a MySQL 8.0
LOAD DATA script. Original workbooks are read only.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl


WORKBOOKS = {
    "PERSON": "30位企业家个人全维度数据采集表.xlsx",
    "FAMILY": "30位企业家家庭维度数据采集表.xlsx",
    "ENTERPRISE": "30位企业家企业全维度数据采集表.xlsx",
    "SOCIAL": "30位企业家社会维度数据采集表.xlsx",
}

CANONICAL_HEADER_ALIASES = {"、": "序号"}


def json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if value is None:
        return None
    return value


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def required_index(headers: list[str], header: str) -> int:
    try:
        return headers.index(header)
    except ValueError as error:
        raise ValueError(f"缺少用于跨表关联的列：{header}") from error


def export_workbook(
    dimension: str,
    source_path: Path,
    staging_writer: csv.DictWriter,
    evidence_writer: csv.DictWriter,
) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(source_path, read_only=True, data_only=False)
    worksheet = workbook.active
    source_rows = worksheet.iter_rows(values_only=True)
    header_row = next(source_rows, None)
    if header_row is None:
        raise ValueError(f"工作簿没有表头：{source_path.name}")

    headers = ["" if cell is None else str(cell).strip() for cell in header_row]
    if any(not item for item in headers):
        raise ValueError(f"工作簿存在空表头：{source_path.name}")
    if len(set(headers)) != len(headers):
        raise ValueError(f"工作簿存在重复表头：{source_path.name}")

    sequence_index = required_index(headers, "序号") if "序号" in headers else required_index(headers, "、")
    person_index = required_index(headers, "企业家")
    enterprise_index = required_index(headers, "核心关联企业")

    row_count = 0
    evidence_count = 0
    for excel_source_row_number, values in enumerate(source_rows, start=2):
        values = list(values)
        if all(value is None or str(value).strip() == "" for value in values):
            continue

        if len(values) != len(headers):
            raise ValueError(
                f"{source_path.name} 第 {excel_source_row_number} 行的列数与表头不一致"
            )

        raw_cells = {headers[index]: json_value(value) for index, value in enumerate(values)}
        person_name = str(values[person_index]).strip() if values[person_index] is not None else ""
        enterprise_name = str(values[enterprise_index]).strip() if values[enterprise_index] is not None else ""
        sequence = values[sequence_index]
        if not person_name or not enterprise_name:
            raise ValueError(f"{source_path.name} 第 {excel_source_row_number} 行缺少企业家或核心关联企业")

        staging_writer.writerow(
            {
                "data_dimension": dimension,
                "source_sequence": sequence,
                "person_name": person_name,
                "core_enterprise_name": enterprise_name,
                "source_file_name": source_path.name,
                "sheet_name": worksheet.title,
                "source_row_number": excel_source_row_number,
                "raw_cells": json.dumps(raw_cells, ensure_ascii=False, separators=(",", ":")),
                "parse_status": "PENDING",
                "parse_message": "",
            }
        )

        for column_index, value in enumerate(values, start=1):
            if value is None or str(value).strip() == "":
                continue
            original_header = headers[column_index - 1]
            evidence_writer.writerow(
                {
                    "file_name": source_path.name,
                    "sheet_name": worksheet.title,
                    "source_row_number": excel_source_row_number,
                    "column_name": original_header,
                    "cell_reference": f"{openpyxl.utils.get_column_letter(column_index)}{excel_source_row_number}",
                    "original_text": json_value(value),
                    "source_level": "S0",
                    "source_date": "",
                    "source_locator": f"{source_path.name}!{worksheet.title}!{openpyxl.utils.get_column_letter(column_index)}{excel_source_row_number}",
                }
            )
            evidence_count += 1
        row_count += 1

    return {
        "dimension": dimension,
        "file_name": source_path.name,
        "sheet_name": worksheet.title,
        "sha256": sha256_file(source_path),
        "record_count": row_count,
        "evidence_cell_count": evidence_count,
        "headers": headers,
    }


def write_loader_sql(output_dir: Path) -> None:
    content = """-- MySQL 8.0 staging loader generated by import_excel_to_staging.py
-- Replace the batch name and operator before executing. LOCAL INFILE must be enabled.
SET NAMES utf8mb4;
START TRANSACTION;

INSERT INTO import_batch (batch_name, source_description, import_status, operator_name)
VALUES ('entrepreneur-four-dimension-v1', '四维企业家受控演示样例 Excel 导入', 'LOADING', 'TO_BE_CONFIRMED');
SET @batch_id = LAST_INSERT_ID();

LOAD DATA LOCAL INFILE 'source_documents.csv'
INTO TABLE source_document
CHARACTER SET utf8mb4
FIELDS TERMINATED BY ',' ENCLOSED BY '"' ESCAPED BY '"'
LINES TERMINATED BY '\\r\\n'
IGNORE 1 LINES
(file_name, sheet_name, source_row_number, column_name, cell_reference, original_text, source_level, @source_date, source_locator)
SET import_batch_id = @batch_id,
    source_date = NULLIF(@source_date, '');

LOAD DATA LOCAL INFILE 'stg_import_row.csv'
INTO TABLE stg_import_row
CHARACTER SET utf8mb4
FIELDS TERMINATED BY ',' ENCLOSED BY '"' ESCAPED BY '"'
LINES TERMINATED BY '\\r\\n'
IGNORE 1 LINES
(data_dimension, source_sequence, person_name, core_enterprise_name, source_file_name, sheet_name, source_row_number, raw_cells, parse_status, @parse_message)
SET import_batch_id = @batch_id,
    parse_message = NULLIF(@parse_message, '');

UPDATE import_batch
SET import_status = 'STAGED',
    record_count = (SELECT COUNT(*) FROM stg_import_row WHERE import_batch_id = @batch_id)
WHERE import_batch_id = @batch_id;
COMMIT;
"""
    (output_dir / "load_staging.sql").write_text(content, encoding="utf-8", newline="\n")


def main() -> None:
    parser = argparse.ArgumentParser(description="Export entrepreneur Excel files to MySQL staging artifacts.")
    parser.add_argument("--input-dir", required=True, type=Path, help="Directory containing the four source workbooks.")
    parser.add_argument("--output-dir", required=True, type=Path, help="Directory for generated staging artifacts.")
    args = parser.parse_args()

    input_dir: Path = args.input_dir
    output_dir: Path = args.output_dir
    if not input_dir.is_dir():
        raise SystemExit(f"输入目录不存在：{input_dir}")
    output_dir.mkdir(parents=True, exist_ok=True)

    staging_fields = [
        "data_dimension", "source_sequence", "person_name", "core_enterprise_name",
        "source_file_name", "sheet_name", "source_row_number", "raw_cells", "parse_status", "parse_message",
    ]
    evidence_fields = [
        "file_name", "sheet_name", "source_row_number", "column_name", "cell_reference",
        "original_text", "source_level", "source_date", "source_locator",
    ]
    manifest: dict[str, Any] = {
        "generated_at": datetime.now().astimezone().isoformat(),
        "source_level": "S0",
        "canonical_header_aliases": CANONICAL_HEADER_ALIASES,
        "workbooks": [],
    }

    with (output_dir / "stg_import_row.csv").open("w", encoding="utf-8", newline="") as staging_file, \
         (output_dir / "source_documents.csv").open("w", encoding="utf-8", newline="") as evidence_file:
        staging_writer = csv.DictWriter(staging_file, fieldnames=staging_fields, quoting=csv.QUOTE_ALL)
        evidence_writer = csv.DictWriter(evidence_file, fieldnames=evidence_fields, quoting=csv.QUOTE_ALL)
        staging_writer.writeheader()
        evidence_writer.writeheader()
        for dimension, file_name in WORKBOOKS.items():
            source_path = input_dir / file_name
            if not source_path.is_file():
                raise SystemExit(f"缺少源工作簿：{source_path}")
            manifest["workbooks"].append(export_workbook(dimension, source_path, staging_writer, evidence_writer))

    manifest["total_staging_rows"] = sum(item["record_count"] for item in manifest["workbooks"])
    manifest["total_evidence_cells"] = sum(item["evidence_cell_count"] for item in manifest["workbooks"])
    (output_dir / "import_manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8", newline="\n"
    )
    write_loader_sql(output_dir)
    print(json.dumps({"output_dir": str(output_dir), "rows": manifest["total_staging_rows"], "evidence_cells": manifest["total_evidence_cells"]}, ensure_ascii=False))


if __name__ == "__main__":
    main()
